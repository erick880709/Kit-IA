"""Pruebas de la Épica E6: gestión de modelos y dashboard (HU-E6-01/02/03)."""

from __future__ import annotations

import json
from datetime import UTC, date, datetime

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.domain.base import Base
from app.domain.entities import EventoTriaje, Modelo, Paciente
from app.domain.exceptions import ValidationError
from app.services import dashboard_service, modelo_service


@pytest.fixture()
def session() -> Session:
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)
    with factory() as s:
        yield s


def _paciente(session, doc: str) -> Paciente:
    p = Paciente(
        tipo_documento="CC", numero_documento=doc, nombres="X", apellidos="Y",
        fecha_nacimiento=date(1990, 1, 1), sexo="Femenino", via_llegada="Ambulancia",
        contacto_emergencia="Z", numero_contacto_emergencia="311",
        departamento="Cundinamarca", ciudad="Bogotá D.C.",
    )
    session.add(p)
    session.commit()
    return p


def _evento(session, *, sugerido: str, asignado: str, motivo=None, doc="1") -> EventoTriaje:
    p = _paciente(session, doc)
    e = EventoTriaje(
        paciente_id=p.id, estado="Cerrado", nivel_sugerido_ia=sugerido,
        nivel_asignado_profesional=asignado,
        concordancia=sugerido == asignado,
        motivo_discrepancia=motivo,
        inicio=datetime.now(UTC) - __import__("datetime").timedelta(minutes=40),
        cierre=datetime.now(UTC),
    )
    session.add(e)
    session.commit()
    return e


# ---------- HU-E6-02 · gestión de modelos ----------

def test_registrar_modelos_primera_activa(session):
    m1 = modelo_service.registrar(
        session, version="v1", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 1),
        ruta_artefacto="modelo-v1.joblib", usuario_id="u1",
    )
    m2 = modelo_service.registrar(
        session, version="v2", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 2),
        ruta_artefacto="modelo-v2.joblib", usuario_id="u1",
    )
    assert m1.activo is True and m2.activo is False  # CA2: nueva queda inactiva


def test_activar_modelo_hace_rollback(session):
    modelo_service.registrar(
        session, version="v1", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 1),
        ruta_artefacto="m1.joblib", usuario_id="u1",
    )
    modelo_service.registrar(
        session, version="v2", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 2),
        ruta_artefacto="m2.joblib", usuario_id="u1",
    )
    activado = modelo_service.activar(session, version="v2", usuario_id="u2")
    assert activado.activo is True
    assert modelo_service.modelo_activo(session).version == "v2"
    # rollback a v1 con un clic
    modelo_service.activar(session, version="v1", usuario_id="u2")
    assert modelo_service.modelo_activo(session).version == "v1"
    historial = modelo_service.historial_activaciones(session)
    acciones = [h.accion for h in historial]
    assert "ACTIVAR_MODELO" in acciones and "REGISTRAR_MODELO" in acciones


def test_activar_modelo_inexistente_rechazado(session):
    with pytest.raises(ValidationError):
        modelo_service.activar(session, version="nope", usuario_id="u1")


def test_desactivar_unico_activo_rechazado(session):
    modelo_service.registrar(
        session, version="v1", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 1),
        ruta_artefacto="m1.joblib", usuario_id="u1",
    )
    with pytest.raises(ValidationError):
        modelo_service.desactivar(session, version="v1", usuario_id="u1")


def test_desactivar_version_no_activa_ok(session):
    modelo_service.registrar(
        session, version="v1", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 1),
        ruta_artefacto="m1.joblib", usuario_id="u1",
    )
    modelo_service.registrar(
        session, version="v2", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 2),
        ruta_artefacto="m2.joblib", usuario_id="u1",
    )
    modelo_service.desactivar(session, version="v2", usuario_id="u1")
    assert modelo_service.modelo_activo(session).version == "v1"


# ---------- HU-E6-01 · dashboard ----------

def test_dashboard_indicadores_ca1_ca3(session):
    _evento(session, sugerido="II", asignado="II")
    _evento(session, sugerido="III", asignado="III", doc="2")
    _evento(
        session, sugerido="II", asignado="I", motivo="riesgo vital", doc="3"
    )
    indicadores = dashboard_service.calcular_indicadores(session)
    assert indicadores["n_eventos"] == 3 and indicadores["n_cerrados"] == 3
    assert abs(sum(indicadores["distribucion"].values()) - 1.0) < 1e-6
    assert indicadores["concordancia_global"] == pytest.approx(2 / 3, abs=1e-3)
    assert indicadores["matriz_confusion"].shape == (5, 5)
    assert len(indicadores["discrepancias"]) == 1
    assert indicadores["discrepancias"][0]["motivo"] == "riesgo vital"
    assert set(indicadores["semaforo"]) == {"f1", "precision", "recall", "auc_roc"}


def test_dashboard_semaforo_con_modelo_activo(session):
    session.add(
        Modelo(
            version="v1", algoritmo="xgb", fecha_entrenamiento=date(2026, 8, 1),
            metricas_json=json.dumps(
                {"macro": {"f1": 0.85, "precision": 0.9, "recall": 0.88},
                 "auc_roc_ovr": 0.96}
            ),
            ruta_artefacto="m1.joblib", activo=True,
        )
    )
    session.commit()
    _evento(session, sugerido="II", asignado="II")
    indicadores = dashboard_service.calcular_indicadores(session)
    assert indicadores["desempeno_ia"]["f1"] == pytest.approx(0.85)
    assert indicadores["semaforo"]["f1"]["estado"] == "ok"
    assert indicadores["semaforo"]["auc_roc"]["estado"] == "ok"


def test_dashboard_sin_eventos_no_falla(session):
    indicadores = dashboard_service.calcular_indicadores(session)
    assert indicadores["n_eventos"] == 0
    assert indicadores["tiempo_promedio_atencion_min"] is None
    assert indicadores["concordancia_global"] is None


# ---------- HU-E6-03 · exportación ----------

def test_exportar_reporte_formatos(session):
    _evento(session, sugerido="III", asignado="III")
    indicadores = dashboard_service.calcular_indicadores(session)
    csv_bytes, nombre = dashboard_service.exportar_reporte(indicadores, formato="csv")
    assert nombre.endswith(".csv") and b"Concordancia" in csv_bytes
    xlsx_bytes, _ = dashboard_service.exportar_reporte(indicadores, formato="excel")
    assert xlsx_bytes[:2] == b"PK"
    pdf_bytes, _ = dashboard_service.exportar_reporte(indicadores, formato="pdf")
    assert pdf_bytes.startswith(b"%PDF")
    with pytest.raises(ValueError):
        dashboard_service.exportar_reporte(indicadores, formato="json")


def test_exportar_excel_incluye_graficos(session):
    """HU-E6-03: el Excel debe bajar CON gráficos (barras + línea de tendencia)."""
    from io import BytesIO

    from openpyxl import load_workbook

    _evento(session, sugerido="III", asignado="III")
    indicadores = dashboard_service.calcular_indicadores(session)
    xlsx_bytes, _ = dashboard_service.exportar_reporte(
        indicadores, formato="excel",
        tendencia=[{"fecha": "2026-08-13", "n": 2}, {"fecha": "2026-08-14", "n": 1}],
    )
    wb = load_workbook(BytesIO(xlsx_bytes))
    assert "graficos" in wb.sheetnames
    assert len(wb["graficos"]._charts) == 3  # 2 barras + 1 línea  # noqa: SLF001


def test_exportar_pdf_incluye_graficos(session):
    """HU-E6-03: el PDF debe bajar CON los gráficos dibujados."""
    _evento(session, sugerido="II", asignado="II")
    indicadores = dashboard_service.calcular_indicadores(session)
    pdf_bytes, _ = dashboard_service.exportar_reporte(
        indicadores, formato="pdf", tendencia=[{"fecha": "2026-08-14", "n": 1}],
    )
    assert pdf_bytes.startswith(b"%PDF")
    assert b"Distribucion de triaje por nivel" in pdf_bytes
    assert b"Concordancia IA vs profesional por nivel" in pdf_bytes
    assert b"Tendencia diaria de eventos" in pdf_bytes


def test_exportar_reporte_sin_identificadores(session):
    p = _paciente(session, "99999999")
    e = EventoTriaje(
        paciente_id=p.id, estado="Cerrado", nivel_sugerido_ia="III",
        nivel_asignado_profesional="III", concordancia=True,
        inicio=datetime.now(UTC), cierre=datetime.now(UTC),
    )
    session.add(e)
    session.commit()
    indicadores = dashboard_service.calcular_indicadores(session)
    pdf_bytes, _ = dashboard_service.exportar_reporte(indicadores, formato="pdf")
    assert b"99999999" not in pdf_bytes  # CA2: sin identificadores directos
